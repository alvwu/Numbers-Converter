import streamlit as st
import pandas as pd
import zipfile
import io
import tempfile
import os
import re
from datetime import date, datetime
from dateutil import parser as date_parser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


NON_DATE_WORDS = {
    'id', 'name', 'first', 'last', 'student', 'exam', 'test', 'quiz',
    'assignment', 'homework', 'participation', 'attendance', 'grade',
    'score', 'total', 'final', 'midterm', 'civics', 'other'
}

MONTH_NAME_PATTERN = (
    r'jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|'
    r'jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|'
    r'nov(?:ember)?|dec(?:ember)?'
)
NONE_OPTION = "(None)"


def normalize_date_column_label(column_name):
    """Normalize a possible date header before detection/parsing."""
    column_str = str(column_name).strip()
    column_str = re.sub(r'\s+', ' ', column_str)
    column_str = re.sub(r'\s*\([^)]*\)\s*$', '', column_str)
    column_str = re.sub(
        r'\s*,?\s*(Mon(?:day)?|Tue(?:sday)?|Wed(?:nesday)?|Thu(?:rsday)?|Fri(?:day)?|Sat(?:urday)?|Sun(?:day)?)\.?\s*$',
        '',
        column_str,
        flags=re.IGNORECASE
    )
    return column_str.strip()


def looks_like_date_label(column_name):
    """Check whether a column header contains a date-like label."""
    if column_name is None:
        return False

    if isinstance(column_name, (datetime, date)):
        return True

    column_str = normalize_date_column_label(column_name)
    if not column_str:
        return False

    column_lower = column_str.lower()
    if column_lower in NON_DATE_WORDS or column_lower.isdigit():
        return False

    if re.search(r'\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b', column_str):
        return True

    if re.search(r'\b\d{1,2}\.\d{1,2}(?:\.\d{2,4})?\b', column_str):
        return True

    if re.search(r'\b\d{4}-\d{1,2}-\d{1,2}\b', column_str):
        return True

    if re.search(rf'\b(?:{MONTH_NAME_PATTERN})\b', column_str, re.IGNORECASE) and re.search(r'\b\d{1,4}\b', column_str):
        return True

    return False


def parse_date_column_value(column_name):
    """Parse a date-like column header into a date object for sorting/export."""
    if column_name is None:
        return None

    if isinstance(column_name, datetime):
        return column_name.date()

    if isinstance(column_name, date):
        return column_name

    column_str = normalize_date_column_label(column_name)
    if not looks_like_date_label(column_str):
        return None

    try:
        parsed = date_parser.parse(column_str, fuzzy=True, dayfirst=False, yearfirst=False)
        if 2000 <= parsed.year <= 2100:
            return parsed.date()
    except (ValueError, TypeError, OverflowError):
        return None

    return None


def is_date_column(column_name):
    """Check if a column name looks like a date."""
    return parse_date_column_value(column_name) is not None


def detect_date_columns(columns, exclude_columns=None):
    """Detect which columns are dates from a list of column names."""
    if exclude_columns is None:
        exclude_columns = []

    date_columns = []
    for col in columns:
        if col not in exclude_columns and is_date_column(col):
            date_columns.append(col)

    return date_columns


def sort_date_columns(columns):
    """Sort selected attendance columns chronologically when they parse as dates."""
    parsed_columns = []
    unparsed_columns = []

    for index, col in enumerate(columns):
        parsed_date = parse_date_column_value(col)
        if parsed_date is None:
            unparsed_columns.append((index, col))
        else:
            parsed_columns.append((parsed_date, index, col))

    parsed_columns.sort(key=lambda item: (item[0], item[1]))

    return [col for _, _, col in parsed_columns] + [col for _, col in unparsed_columns]


def normalize_optional_column(column_name):
    """Convert optional UI selections into usable column names."""
    if column_name in (None, "", NONE_OPTION):
        return None
    return column_name


def get_row_text_value(row, column_name):
    """Return a stripped string value for an optional dataframe column."""
    column_name = normalize_optional_column(column_name)
    if column_name is None:
        return ""

    try:
        value = row[column_name]
    except (KeyError, TypeError):
        return ""

    return str(value).strip() if pd.notna(value) else ""


def build_student_mask(df, column_names):
    """Build a mask for rows that have data in at least one selected identifier column."""
    selected_columns = [normalize_optional_column(col) for col in column_names]
    selected_columns = [col for col in selected_columns if col]

    if not selected_columns:
        return pd.Series(False, index=df.index)

    mask = pd.Series(False, index=df.index)
    for col in selected_columns:
        mask = mask | (df[col].notna() & (df[col].astype(str).str.strip() != ''))

    return mask


def parse_numbers_file(uploaded_file):
    """Parse a .numbers file and return a pandas DataFrame."""
    with tempfile.TemporaryDirectory() as temp_dir:
        # Save uploaded file to temp location
        temp_numbers_path = os.path.join(temp_dir, "temp.numbers")
        with open(temp_numbers_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # .numbers files are zip archives — guard against zip bombs
        MAX_EXTRACTED_BYTES = 100 * 1024 * 1024  # 100 MB
        with zipfile.ZipFile(temp_numbers_path, 'r') as zip_ref:
            total_size = 0
            for info in zip_ref.infolist():
                total_size += info.file_size
                if total_size > MAX_EXTRACTED_BYTES:
                    raise ValueError("Uploaded file expands to more than 100 MB when decompressed.")
                zip_ref.extract(info, temp_dir)

        # Find and parse the tables
        tables_dir = os.path.join(temp_dir, "Index", "Tables")

        if not os.path.exists(tables_dir):
            # Try alternative structure
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    if file.endswith('.iwa'):
                        pass  # .iwa files need special parsing
            raise ValueError("Could not find tables in the Numbers file. Please try exporting as CSV.")

        # For .numbers files, we need to use the numbers-parser library
        try:
            from numbers_parser import Document
            doc = Document(temp_numbers_path)
            sheets = doc.sheets
            if not sheets:
                raise ValueError("No sheets found in the Numbers file.")

            # Get the first table from the first sheet
            table = sheets[0].tables[0]

            # Extract data
            data = []
            headers = []

            for row_num, row in enumerate(table.iter_rows()):
                row_data = []
                for cell in row:
                    row_data.append(cell.value if cell.value is not None else "")

                if row_num == 0:
                    headers = row_data
                else:
                    data.append(row_data)

            # Convert headers to strings and handle datetime objects
            processed_headers = []
            for i, h in enumerate(headers):
                if h is None or h == "":
                    # Generate unique name for empty columns
                    processed_headers.append(f"Column_{i+1}")
                elif isinstance(h, datetime):
                    # Convert datetime to string format
                    processed_headers.append(h.strftime("%m/%d/%Y"))
                else:
                    processed_headers.append(str(h))

            # Handle duplicate column names by appending a suffix
            seen = {}
            unique_headers = []
            for h in processed_headers:
                if h in seen:
                    seen[h] += 1
                    unique_headers.append(f"{h}_{seen[h]}")
                else:
                    seen[h] = 0
                    unique_headers.append(h)

            df = pd.DataFrame(data, columns=unique_headers)
            return df

        except ImportError:
            raise ImportError("The 'numbers-parser' library is required. Please install it with: pip install numbers-parser")


def categorize_columns(columns, category_keywords):
    """Categorize columns based on keywords."""
    categorized = {}
    uncategorized = []

    for col in columns:
        col_lower = col.lower()
        found_category = None

        for category, keywords in category_keywords.items():
            for keyword in keywords:
                if keyword.lower() in col_lower:
                    found_category = category
                    break
            if found_category:
                break

        if found_category:
            if found_category not in categorized:
                categorized[found_category] = []
            categorized[found_category].append(col)
        else:
            uncategorized.append(col)

    return categorized, uncategorized


def create_student_excel(df, id_column, first_name_column, last_name_column, category_keywords, show_category_averages, category_max_points=None, category_weights=None, item_max_points=None):
    """Create an Excel file with each student on their own sheet."""
    output = io.BytesIO()
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Columns to exclude from grade columns (student identifier columns)
    identifier_columns = [id_column, first_name_column, last_name_column]

    # Get grade columns (all columns except the identifier columns)
    grade_columns = [col for col in df.columns if col not in identifier_columns]

    # Categorize columns
    categorized, uncategorized = categorize_columns(grade_columns, category_keywords)

    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    category_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    category_font = Font(bold=True, size=11)
    weight_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange for weighted section
    excused_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Light orange/yellow for excused
    zero_score_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red for zero scores
    score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for scores above zero
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Initialize category_max_points if not provided
    if category_max_points is None:
        category_max_points = {}

    # Initialize category_weights if not provided
    if category_weights is None:
        category_weights = {}

    # Initialize item_max_points if not provided
    if item_max_points is None:
        item_max_points = {}

    # Sort dataframe by last name alphabetically
    df_sorted = df.copy()
    df_sorted['_sort_key'] = df_sorted[last_name_column].astype(str).str.lower()
    df_sorted = df_sorted.sort_values('_sort_key')
    df_sorted = df_sorted.drop('_sort_key', axis=1)

    # Create a sheet for each student
    for idx, row in df_sorted.iterrows():
        student_id = str(row[id_column]).strip() if pd.notna(row[id_column]) else ""
        first_name = str(row[first_name_column]).strip() if pd.notna(row[first_name_column]) else ""
        last_name = str(row[last_name_column]).strip() if pd.notna(row[last_name_column]) else ""

        # Skip rows where ID, first name, and last name are all empty
        if not student_id and not first_name and not last_name:
            continue

        # Create sheet name as "Last Name, First Name"
        sheet_name = f"{last_name}, {first_name}"

        # Sanitize sheet name (Excel has restrictions)
        safe_name = sheet_name[:31]  # Max 31 chars
        safe_name = ''.join(c for c in safe_name if c not in '[]:*?/\\')
        if not safe_name or safe_name == ", ":
            safe_name = f"Student_{idx}"

        # Handle duplicate sheet names
        original_name = safe_name
        counter = 1
        while safe_name in wb.sheetnames:
            safe_name = f"{original_name[:28]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=safe_name)

        # Add student identifier info at top of sheet
        ws['A1'] = "ID:"
        ws['B1'] = student_id
        ws['A1'].font = header_font
        ws['B1'].font = Font(bold=True, size=12)

        ws['A2'] = "First Name:"
        ws['B2'] = first_name
        ws['A2'].font = header_font
        ws['B2'].font = Font(bold=True, size=12)

        ws['A3'] = "Last Name:"
        ws['B3'] = last_name
        ws['A3'].font = header_font
        ws['B3'].font = Font(bold=True, size=12)

        current_row = 5  # Start after ID, First Name, Last Name, and a blank row
        all_grades = []
        all_max_points = []
        category_averages = {}
        total_excused = 0  # Track total excused assignments

        # Add headers
        ws.cell(row=current_row, column=1, value="Categories")
        ws.cell(row=current_row, column=2, value="Score")
        ws.cell(row=current_row, column=3, value="Max Points")
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).font = header_font_white
            ws.cell(row=current_row, column=col).fill = header_fill
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = center_align

        current_row += 1

        # Add categorized grades
        for category, columns in categorized.items():
            # Category header
            ws.cell(row=current_row, column=1, value=category.upper())
            ws.cell(row=current_row, column=1).font = category_font
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).fill = category_fill
                ws.cell(row=current_row, column=col).border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1

            # Get default max points for this category
            default_max_points = category_max_points.get(category, 100)

            category_grades = []
            category_max = []
            for col in columns:
                cell_value = row[col]
                is_excused = False

                # Check if the grade is excused (E or e)
                if pd.notna(cell_value) and str(cell_value).strip().upper() == 'E':
                    is_excused = True
                    grade = 'E'
                    total_excused += 1
                else:
                    try:
                        grade = float(cell_value) if pd.notna(cell_value) and cell_value != '' else 0
                    except (ValueError, TypeError):
                        grade = 0

                # Use per-item max points if set, otherwise use category default
                max_points = item_max_points.get(col, default_max_points)

                ws.cell(row=current_row, column=1, value=col)
                ws.cell(row=current_row, column=2, value=grade)
                ws.cell(row=current_row, column=3, value=max_points if not is_excused else "Excused")
                for c in range(1, 4):
                    ws.cell(row=current_row, column=c).border = border
                ws.cell(row=current_row, column=2).alignment = center_align
                ws.cell(row=current_row, column=3).alignment = center_align

                # Apply styling based on grade status
                if is_excused:
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = excused_fill
                elif grade == 0:
                    # Highlight zero scores with light red
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = zero_score_fill
                    # Still count zero scores toward averages
                    category_grades.append(grade)
                    category_max.append(max_points)
                    all_grades.append(grade)
                    all_max_points.append(max_points)
                else:
                    # Highlight scores above zero with light green
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = score_fill
                    # Count non-excused grades toward averages
                    category_grades.append(grade)
                    category_max.append(max_points)
                    all_grades.append(grade)
                    all_max_points.append(max_points)

                current_row += 1

            # Calculate category average as percentage
            if category_grades and category_max:
                total_earned = sum(category_grades)
                total_possible = sum(category_max)
                category_averages[category] = (total_earned / total_possible * 100) if total_possible > 0 else 0

        # Add uncategorized grades
        if uncategorized:
            ws.cell(row=current_row, column=1, value="OTHER")
            ws.cell(row=current_row, column=1).font = category_font
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).fill = category_fill
                ws.cell(row=current_row, column=col).border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1

            # Get default max points for "Other" category
            default_other_max_points = category_max_points.get("Other", 100)

            other_grades = []
            other_max = []
            for col in uncategorized:
                cell_value = row[col]
                is_excused = False

                # Check if the grade is excused (E or e)
                if pd.notna(cell_value) and str(cell_value).strip().upper() == 'E':
                    is_excused = True
                    grade = 'E'
                    total_excused += 1
                else:
                    try:
                        grade = float(cell_value) if pd.notna(cell_value) and cell_value != '' else 0
                    except (ValueError, TypeError):
                        grade = 0

                # Use per-item max points if set, otherwise use "Other" default
                other_max_points = item_max_points.get(col, default_other_max_points)

                ws.cell(row=current_row, column=1, value=col)
                ws.cell(row=current_row, column=2, value=grade)
                ws.cell(row=current_row, column=3, value=other_max_points if not is_excused else "Excused")
                for c in range(1, 4):
                    ws.cell(row=current_row, column=c).border = border
                ws.cell(row=current_row, column=2).alignment = center_align
                ws.cell(row=current_row, column=3).alignment = center_align

                # Apply styling based on grade status
                if is_excused:
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = excused_fill
                elif grade == 0:
                    # Highlight zero scores with light red
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = zero_score_fill
                    # Still count zero scores toward averages
                    other_grades.append(grade)
                    other_max.append(other_max_points)
                    all_grades.append(grade)
                    all_max_points.append(other_max_points)
                else:
                    # Highlight scores above zero with light green
                    for c in range(1, 4):
                        ws.cell(row=current_row, column=c).fill = score_fill
                    # Count non-excused grades toward averages
                    other_grades.append(grade)
                    other_max.append(other_max_points)
                    all_grades.append(grade)
                    all_max_points.append(other_max_points)

                current_row += 1

            if other_grades and other_max:
                total_earned = sum(other_grades)
                total_possible = sum(other_max)
                category_averages["Other"] = (total_earned / total_possible * 100) if total_possible > 0 else 0

        current_row += 1

        # Add category averages if enabled
        if show_category_averages and category_averages:
            ws.cell(row=current_row, column=1, value="CATEGORY AVERAGES (%)")
            ws.cell(row=current_row, column=1).font = header_font_white
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).fill = header_fill
                ws.cell(row=current_row, column=col).border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            current_row += 1

            for category, avg in category_averages.items():
                ws.cell(row=current_row, column=1, value=category)
                ws.cell(row=current_row, column=2, value=f"{round(avg, 2)}%")
                for c in range(1, 4):
                    ws.cell(row=current_row, column=c).border = border
                ws.cell(row=current_row, column=2).alignment = center_align
                current_row += 1

            current_row += 1

        # Add weighted grades section
        if category_weights and category_averages:
            # Header for weighted grades section
            ws.cell(row=current_row, column=1, value="WEIGHTED GRADES")
            ws.cell(row=current_row, column=1).font = header_font_white
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).fill = header_fill
                ws.cell(row=current_row, column=col).border = border
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

            # Column headers for weighted section
            ws.cell(row=current_row, column=1, value="Category")
            ws.cell(row=current_row, column=2, value="Score (%)")
            ws.cell(row=current_row, column=3, value="Weight (%)")
            ws.cell(row=current_row, column=4, value="Weighted Score")
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).font = category_font
                ws.cell(row=current_row, column=col).fill = weight_fill
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).alignment = center_align
            current_row += 1

            # Calculate weighted scores for each category
            total_weighted_score = 0
            total_weight_used = 0

            for category, avg_percentage in category_averages.items():
                weight = category_weights.get(category, 0)
                weighted_score = (avg_percentage * weight) / 100 if weight > 0 else 0

                ws.cell(row=current_row, column=1, value=category)
                ws.cell(row=current_row, column=2, value=f"{round(avg_percentage, 2)}%")
                ws.cell(row=current_row, column=3, value=f"{weight}%")
                ws.cell(row=current_row, column=4, value=round(weighted_score, 2))
                for col in range(1, 5):
                    ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=2).alignment = center_align
                ws.cell(row=current_row, column=3).alignment = center_align
                ws.cell(row=current_row, column=4).alignment = center_align

                total_weighted_score += weighted_score
                total_weight_used += weight
                current_row += 1

            current_row += 1

            # Final weighted grade
            ws.cell(row=current_row, column=1, value="FINAL WEIGHTED GRADE")
            ws.cell(row=current_row, column=2, value=f"{round(total_weighted_score, 2)}%")
            ws.cell(row=current_row, column=3, value=f"(of {total_weight_used}%)")
            for col in range(1, 5):
                ws.cell(row=current_row, column=col).font = Font(bold=True, size=12)
                ws.cell(row=current_row, column=col).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=3).alignment = center_align

        # Add excused summary if there are any excused assignments
        if total_excused > 0:
            current_row += 2

            ws.cell(row=current_row, column=1, value="EXCUSED ASSIGNMENTS")
            ws.cell(row=current_row, column=2, value=total_excused)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=current_row, column=2).font = Font(bold=True, size=11)
            for col in range(1, 4):
                ws.cell(row=current_row, column=col).fill = excused_fill
                ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=2).alignment = center_align

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    wb.save(output)
    output.seek(0)
    return output


def create_attendance_excel(df, id_column, first_name_column, last_name_column, attendance_columns):
    """Create an Excel file with attendance records for each student."""
    output = io.BytesIO()
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Color fills for attendance grades
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for 1 (present)
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red/orange for 0 (absent)

    # Sort dataframe by last name alphabetically
    df_sorted = df.copy()
    sort_column = next(
        (col for col in [normalize_optional_column(last_name_column), normalize_optional_column(first_name_column), normalize_optional_column(id_column)] if col),
        None
    )
    if sort_column:
        df_sorted['_sort_key'] = df_sorted[sort_column].fillna('').astype(str).str.lower()
    else:
        df_sorted['_sort_key'] = df_sorted.index.astype(str)
    df_sorted = df_sorted.sort_values('_sort_key')
    df_sorted = df_sorted.drop('_sort_key', axis=1)
    sorted_attendance_columns = sort_date_columns(attendance_columns)

    # Create a sheet for each student
    for idx, row in df_sorted.iterrows():
        student_id = get_row_text_value(row, id_column)
        first_name = get_row_text_value(row, first_name_column)
        last_name = get_row_text_value(row, last_name_column)

        # Skip rows where ID, first name, and last name are all empty
        if not student_id and not first_name and not last_name:
            continue

        # Create sheet name from the best available identifier
        if last_name and first_name:
            sheet_name = f"{last_name}, {first_name}"
        elif last_name:
            sheet_name = last_name
        elif first_name:
            sheet_name = first_name
        elif student_id:
            sheet_name = f"Student {student_id}"
        else:
            sheet_name = f"Student_{idx}"

        # Sanitize sheet name (Excel has restrictions)
        safe_name = sheet_name[:31]  # Max 31 chars
        safe_name = ''.join(c for c in safe_name if c not in '[]:*?/\\')
        if not safe_name or safe_name == ", ":
            safe_name = f"Student_{idx}"

        # Handle duplicate sheet names
        original_name = safe_name
        counter = 1
        while safe_name in wb.sheetnames:
            safe_name = f"{original_name[:28]}_{counter}"
            counter += 1

        ws = wb.create_sheet(title=safe_name)

        # Add student identifier info at top of sheet
        ws['A1'] = "ID:"
        ws['B1'] = student_id
        ws['A1'].font = header_font
        ws['B1'].font = Font(bold=True, size=12)

        ws['A2'] = "First Name:"
        ws['B2'] = first_name
        ws['A2'].font = header_font
        ws['B2'].font = Font(bold=True, size=12)

        ws['A3'] = "Last Name:"
        ws['B3'] = last_name
        ws['A3'].font = header_font
        ws['B3'].font = Font(bold=True, size=12)

        current_row = 5  # Start after student info and a blank row

        # Add headers for attendance
        ws.cell(row=current_row, column=1, value="Date")
        ws.cell(row=current_row, column=2, value="Attendance")
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).font = header_font_white
            ws.cell(row=current_row, column=col).fill = header_fill
            ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=col).alignment = center_align

        current_row += 1

        # Add attendance records
        total_present = 0
        total_days = 0

        for date_col in sorted_attendance_columns:
            try:
                grade = float(row[date_col]) if pd.notna(row[date_col]) and row[date_col] != '' else 0
                grade = int(grade) if grade in [0, 1] else grade
            except (ValueError, TypeError):
                grade = 0

            parsed_date = parse_date_column_value(date_col)
            date_cell = ws.cell(
                row=current_row,
                column=1,
                value=parsed_date if parsed_date is not None else str(date_col)
            )
            if parsed_date is not None:
                date_cell.number_format = "mm/dd/yyyy"
            ws.cell(row=current_row, column=2, value=grade)

            # Apply styling
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=2).border = border
            ws.cell(row=current_row, column=2).alignment = center_align

            # Color code based on attendance (0 = absent/orange, 1 = present/green)
            if grade == 1:
                ws.cell(row=current_row, column=2).fill = present_fill
                total_present += 1
            elif grade == 0:
                ws.cell(row=current_row, column=2).fill = absent_fill

            total_days += 1
            current_row += 1

        # Add summary row
        current_row += 1
        attendance_rate = (total_present / total_days * 100) if total_days > 0 else 0

        ws.cell(row=current_row, column=1, value="ATTENDANCE SUMMARY")
        ws.cell(row=current_row, column=1).font = header_font_white
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=2).fill = header_fill
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2).border = border
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        current_row += 1

        ws.cell(row=current_row, column=1, value="Days Present:")
        ws.cell(row=current_row, column=2, value=total_present)
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        total_absent = total_days - total_present
        ws.cell(row=current_row, column=1, value="Days Absent:")
        ws.cell(row=current_row, column=2, value=total_absent)
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        ws.cell(row=current_row, column=1, value="Total Days:")
        ws.cell(row=current_row, column=2, value=total_days)
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = center_align
        current_row += 1

        ws.cell(row=current_row, column=1, value="Attendance Rate:")
        ws.cell(row=current_row, column=2, value=f"{round(attendance_rate, 1)}%")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.cell(row=current_row, column=2).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.cell(row=current_row, column=1).border = border
        ws.cell(row=current_row, column=2).border = border
        ws.cell(row=current_row, column=2).alignment = center_align

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    wb.save(output)
    output.seek(0)
    return output


def main():
    st.set_page_config(
        page_title="GradeBook Transfer",
        page_icon="📚",
        layout="wide"
    )

    # Custom CSS for better styling
    st.markdown("""
        <style>
        .main-header {
            font-size: 2.5rem;
            font-weight: bold;
            color: #1E3A5F;
            text-align: center;
            margin-bottom: 0.5rem;
        }
        .sub-header {
            font-size: 1.1rem;
            color: #666;
            text-align: center;
            margin-bottom: 2rem;
        }
        .success-box {
            padding: 1rem;
            border-radius: 0.5rem;
            background-color: #d4edda;
            border: 1px solid #c3e6cb;
            margin: 1rem 0;
            color: #155724;
        }
        .success-box h4 {
            color: #155724;
        }
        .success-box p, .success-box li {
            color: #155724;
        }
        .info-box {
            padding: 1rem;
            border-radius: 0.5rem;
            background-color: #e7f3ff;
            border: 1px solid #b6d4fe;
            margin: 1rem 0;
        }
        .stButton>button {
            width: 100%;
        }
        </style>
    """, unsafe_allow_html=True)

    # Header
    st.markdown('<p class="main-header">📚 GradeBook Transfer</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Convert your Numbers gradebook into organized Excel files with individual student sheets</p>', unsafe_allow_html=True)

    st.divider()

    # Create tabs
    tab1, tab2 = st.tabs(["📊 Grade Transfer", "📅 Attendance"])

    # Sidebar for settings (shared)
    with st.sidebar:
        st.header("⚙️ Grade Transfer Settings")

        st.subheader("Grade Display Options")
        show_category_averages = st.checkbox(
            "Show category averages",
            value=False,
            help="Display average scores for each category (Exams, Assignments, etc.) in addition to the overall final grade"
        )

        st.divider()

        st.subheader("Category Keywords")
        st.caption("Add keywords to categorize grade columns. Columns containing these keywords will be grouped together.")

        # Initialize session state for categories
        if 'categories' not in st.session_state:
            st.session_state.categories = {
                "Exams": ["exam", "test", "midterm", "final"],
                "Assignments": ["assignment", "homework", "hw"],
                "Participation": ["participation", "attendance"],
                "El Civics": ["el civics", "civics", "elcivics"]
            }

        # Initialize session state for category max points (default for each category)
        if 'category_max_points' not in st.session_state:
            st.session_state.category_max_points = {
                "Exams": 100,
                "Assignments": 100,
                "Participation": 1,
                "El Civics": 100,
                "Other": 100
            }

        # Initialize session state for per-item custom max points
        if 'item_max_points' not in st.session_state:
            st.session_state.item_max_points = {}

        # Initialize session state for category weights (percentages)
        if 'category_weights' not in st.session_state:
            st.session_state.category_weights = {
                "Exams": 25,
                "Assignments": 25,
                "Participation": 30,
                "El Civics": 20,
                "Other": 0
            }

        # Display existing categories
        categories_to_remove = []
        for category in list(st.session_state.categories.keys()):
            with st.expander(f"📁 {category}", expanded=False):
                keywords = st.session_state.categories[category]
                new_keywords = st.text_area(
                    "Keywords (one per line)",
                    value="\n".join(keywords),
                    key=f"keywords_{category}",
                    height=100
                )
                st.session_state.categories[category] = [k.strip() for k in new_keywords.split("\n") if k.strip()]

                # Max points setting for this category
                current_max = st.session_state.category_max_points.get(category, 100)
                new_max = st.number_input(
                    "Max points per item",
                    min_value=1,
                    value=current_max,
                    key=f"max_points_{category}",
                    help=f"Each item in {category} is out of this many points"
                )
                st.session_state.category_max_points[category] = new_max

                # Weight percentage for this category
                current_weight = st.session_state.category_weights.get(category, 0)
                new_weight = st.number_input(
                    "Weight (%)",
                    min_value=0,
                    max_value=100,
                    value=current_weight,
                    key=f"weight_{category}",
                    help=f"Weight of {category} in final grade calculation (all weights should sum to 100%)"
                )
                st.session_state.category_weights[category] = new_weight

                if st.button(f"🗑️ Remove {category}", key=f"remove_{category}"):
                    categories_to_remove.append(category)

        # Remove marked categories
        for cat in categories_to_remove:
            del st.session_state.categories[cat]
            if cat in st.session_state.category_max_points:
                del st.session_state.category_max_points[cat]
            if cat in st.session_state.category_weights:
                del st.session_state.category_weights[cat]
            st.rerun()

        # Other category max points (for uncategorized items)
        with st.expander("📁 Other (uncategorized)", expanded=False):
            other_max = st.number_input(
                "Max points per item",
                min_value=1,
                value=st.session_state.category_max_points.get("Other", 100),
                key="max_points_Other",
                help="Each uncategorized item is out of this many points"
            )
            st.session_state.category_max_points["Other"] = other_max

            other_weight = st.number_input(
                "Weight (%)",
                min_value=0,
                max_value=100,
                value=st.session_state.category_weights.get("Other", 0),
                key="weight_Other",
                help="Weight of uncategorized items in final grade calculation"
            )
            st.session_state.category_weights["Other"] = other_weight

        # Show total weight
        total_weight = sum(st.session_state.category_weights.values())
        if total_weight == 100:
            st.success(f"Total weight: {total_weight}%")
        elif total_weight < 100:
            st.warning(f"Total weight: {total_weight}% (should be 100%)")
        else:
            st.error(f"Total weight: {total_weight}% (exceeds 100%)")

        st.divider()

        # Add new category
        st.subheader("Add New Category")
        new_category_name = st.text_input("Category name", placeholder="e.g., Projects")
        new_category_keywords = st.text_input("Keywords (comma-separated)", placeholder="e.g., project, proj")
        new_category_max_points = st.number_input("Max points per item", min_value=1, value=100, key="new_category_max")
        new_category_weight = st.number_input("Weight (%)", min_value=0, max_value=100, value=0, key="new_category_weight")

        if st.button("➕ Add Category"):
            if new_category_name and new_category_keywords:
                keywords_list = [k.strip() for k in new_category_keywords.split(",") if k.strip()]
                if keywords_list:
                    st.session_state.categories[new_category_name] = keywords_list
                    st.session_state.category_max_points[new_category_name] = new_category_max_points
                    st.session_state.category_weights[new_category_name] = new_category_weight
                    st.success(f"Added category: {new_category_name}")
                    st.rerun()
            else:
                st.warning("Please enter both category name and keywords")

    # ==================== GRADE TRANSFER TAB ====================
    with tab1:
        # Main content area
        col1, col2 = st.columns([2, 1])

        with col1:
            st.subheader("📤 Upload Your Numbers File")

            uploaded_file = st.file_uploader(
                "Drag and drop your .numbers file here",
                type=["numbers"],
                help="Upload your Numbers gradebook file. The first row should contain headers (student name and grade categories).",
                key="grades_file_uploader"
            )

        with col2:
            st.subheader("📋 How It Works")
            st.markdown("""
            1. **Upload** your .numbers gradebook
            2. **Select** ID, First Name, and Last Name columns
            3. **Review** the detected categories
            4. **Download** the organized Excel file
            """)

        if uploaded_file is not None:
            st.divider()

            with st.spinner("🔄 Parsing Numbers file..."):
                try:
                    df = parse_numbers_file(uploaded_file)
                    st.success("✅ File parsed successfully!")

                    # Show preview
                    st.subheader("📊 Data Preview")
                    st.dataframe(df.head(10), use_container_width=True)

                    st.info(f"📈 Found **{len(df)}** rows and **{len(df.columns)}** columns")

                    # Select student identifier columns
                    st.subheader("🏷️ Select Student Identifier Columns")
                    st.caption("Select the columns that contain student ID, first name, and last name. These will appear at the top of each student's sheet and will not be categorized as grades.")

                    col_options = df.columns.tolist()

                    # Auto-detect column indices based on column names
                    def find_column_index(columns, keywords, default=0):
                        for i, col in enumerate(columns):
                            col_lower = str(col).lower()
                            for keyword in keywords:
                                if keyword in col_lower:
                                    return i
                        return default

                    id_index = find_column_index(col_options, ['id', 'student id', 'studentid'], 0)
                    lname_index = find_column_index(col_options, ['last name', 'lastname', 'last_name', 'lname'], 1 if len(col_options) > 1 else 0)
                    fname_index = find_column_index(col_options, ['first name', 'firstname', 'first_name', 'fname'], 2 if len(col_options) > 2 else 0)

                    id_col, lname_col, fname_col = st.columns(3)
                    with id_col:
                        id_column = st.selectbox(
                            "ID Column",
                            options=col_options,
                            index=id_index,
                            key="grades_id_col"
                        )
                    with lname_col:
                        last_name_column = st.selectbox(
                            "Last Name Column",
                            options=col_options,
                            index=lname_index,
                            key="grades_lname_col"
                        )
                    with fname_col:
                        first_name_column = st.selectbox(
                            "First Name Column",
                            options=col_options,
                            index=fname_index,
                            key="grades_fname_col"
                        )

                    # Count valid students (rows where at least one identifier is not empty)
                    identifier_columns = [id_column, first_name_column, last_name_column]
                    valid_students = df[
                        (df[id_column].notna() & (df[id_column] != '')) |
                        (df[first_name_column].notna() & (df[first_name_column] != '')) |
                        (df[last_name_column].notna() & (df[last_name_column] != ''))
                    ]
                    st.info(f"📈 Found **{len(valid_students)}** valid students (excluding empty rows)")

                    # Show category detection preview
                    st.subheader("🗂️ Category Detection Preview")
                    grade_columns = [col for col in df.columns if col not in identifier_columns]
                    categorized, uncategorized = categorize_columns(grade_columns, st.session_state.categories)

                    col1, col2 = st.columns(2)

                    with col1:
                        st.markdown("**Categorized Columns:**")
                        for category, columns in categorized.items():
                            with st.expander(f"{category} ({len(columns)} items)"):
                                for col in columns:
                                    st.write(f"• {col}")

                    with col2:
                        st.markdown("**Uncategorized Columns:**")
                        if uncategorized:
                            for col in uncategorized:
                                st.write(f"• {col}")
                            st.caption("💡 Add keywords in the sidebar to categorize these columns")
                        else:
                            st.write("All columns are categorized!")

                    # Custom max points per item
                    st.subheader("🎯 Custom Max Points Per Item")
                    st.caption("Override the default max points for specific assignments. Leave blank to use the category default.")

                    with st.expander("Customize individual item max points", expanded=False):
                        for category, columns in categorized.items():
                            default_max = st.session_state.category_max_points.get(category, 100)
                            st.markdown(f"**{category}** (default: {default_max} pts)")
                            for col in columns:
                                col_key = f"item_max_{col}"
                                current_val = st.session_state.item_max_points.get(col)
                                custom_max = st.number_input(
                                    f"{col}",
                                    min_value=1,
                                    value=current_val if current_val else default_max,
                                    key=col_key,
                                    help=f"Custom max points for {col}"
                                )
                                # Only store if different from default
                                if custom_max != default_max:
                                    st.session_state.item_max_points[col] = custom_max
                                elif col in st.session_state.item_max_points:
                                    del st.session_state.item_max_points[col]

                        if uncategorized:
                            default_other = st.session_state.category_max_points.get("Other", 100)
                            st.markdown(f"**Other** (default: {default_other} pts)")
                            for col in uncategorized:
                                col_key = f"item_max_{col}"
                                current_val = st.session_state.item_max_points.get(col)
                                custom_max = st.number_input(
                                    f"{col}",
                                    min_value=1,
                                    value=current_val if current_val else default_other,
                                    key=col_key,
                                    help=f"Custom max points for {col}"
                                )
                                if custom_max != default_other:
                                    st.session_state.item_max_points[col] = custom_max
                                elif col in st.session_state.item_max_points:
                                    del st.session_state.item_max_points[col]

                        if st.session_state.item_max_points:
                            st.info(f"📝 {len(st.session_state.item_max_points)} item(s) have custom max points")

                    st.divider()

                    # Generate Excel
                    st.subheader("📥 Generate Excel File")

                    if st.button("🚀 Generate Excel File", type="primary", use_container_width=True, key="generate_grades"):
                        with st.spinner("🔄 Creating Excel file with individual student sheets..."):
                            excel_output = create_student_excel(
                                df,
                                id_column,
                                first_name_column,
                                last_name_column,
                                st.session_state.categories,
                                show_category_averages,
                                st.session_state.category_max_points,
                                st.session_state.category_weights,
                                st.session_state.item_max_points
                            )

                            st.success("✅ Excel file generated successfully!")

                            # Summary
                            st.subheader("✨ Generation Complete!")
                            st.write(f"Created **{len(valid_students)}** individual student sheets")
                            st.write("Each sheet contains:")
                            summary_items = [
                                "Student ID, First Name, and Last Name",
                                "All assignments organized by category",
                                "Scores with max points"
                            ]
                            if show_category_averages:
                                summary_items.append("Category averages (%)")
                            summary_items.append("Weighted grades breakdown by category")
                            summary_items.append("Final weighted grade")
                            for item in summary_items:
                                st.write(f"- {item}")

                            # Download button
                            st.download_button(
                                label="📥 Download Excel File",
                                data=excel_output,
                                file_name="gradebook_transfer.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key="download_grades"
                            )

                except ImportError as e:
                    st.error(f"❌ {str(e)}")
                    st.info("💡 Run this command in your terminal: `pip install numbers-parser`")
                except Exception as e:
                    st.error(f"❌ Error parsing file: {str(e)}")
                    st.info("💡 Make sure your Numbers file has a table with headers in the first row")

    # ==================== ATTENDANCE TAB ====================
    with tab2:
        col1, col2 = st.columns([2, 1])

        with col1:
            st.subheader("📤 Upload Your Numbers File")

            attendance_file = st.file_uploader(
                "Drag and drop your .numbers file here",
                type=["numbers"],
                help="Upload your Numbers attendance file. The first row should contain headers (student info and dates).",
                key="attendance_file_uploader"
            )

        with col2:
            st.subheader("📋 How It Works")
            st.markdown("""
            1. **Upload** your .numbers attendance file
            2. **Select** ID, First Name, and Last Name columns
            3. **Select** which columns contain attendance dates
            4. **Download** the organized attendance Excel file

            **Color Coding:**
            - 🟢 Green = Present (1)
            - 🟠 Orange = Absent (0)
            """)

        if attendance_file is not None:
            st.divider()

            with st.spinner("🔄 Parsing Numbers file..."):
                try:
                    att_df = parse_numbers_file(attendance_file)
                    st.success("✅ File parsed successfully!")

                    # Show preview
                    st.subheader("📊 Data Preview")
                    st.dataframe(att_df.head(10), use_container_width=True)

                    st.info(f"📈 Found **{len(att_df)}** rows and **{len(att_df.columns)}** columns")

                    # Select student identifier columns
                    st.subheader("🏷️ Select Student Identifier Columns")

                    att_col_options = att_df.columns.tolist()

                    # Auto-detect column indices based on column names
                    def find_column_index(columns, keywords):
                        for i, col in enumerate(columns):
                            col_lower = str(col).lower()
                            for keyword in keywords:
                                if keyword in col_lower:
                                    return i + 1
                        return 0

                    att_column_options = [NONE_OPTION] + att_col_options
                    att_id_index = find_column_index(att_col_options, ['id', 'student id', 'studentid'])
                    att_lname_index = find_column_index(att_col_options, ['last name', 'lastname', 'last_name', 'lname'])
                    att_fname_index = find_column_index(att_col_options, ['first name', 'firstname', 'first_name', 'fname'])

                    id_col, lname_col, fname_col = st.columns(3)
                    with id_col:
                        att_id_column = st.selectbox(
                            "ID Column",
                            options=att_column_options,
                            index=att_id_index,
                            key="att_id_col"
                        )
                    with lname_col:
                        att_last_name_column = st.selectbox(
                            "Last Name Column",
                            options=att_column_options,
                            index=att_lname_index,
                            key="att_lname_col"
                        )
                    with fname_col:
                        att_first_name_column = st.selectbox(
                            "First Name Column",
                            options=att_column_options,
                            index=att_fname_index,
                            key="att_fname_col"
                        )

                    att_id_column = normalize_optional_column(att_id_column)
                    att_last_name_column = normalize_optional_column(att_last_name_column)
                    att_first_name_column = normalize_optional_column(att_first_name_column)

                    # Count valid students
                    att_identifier_columns = [col for col in [att_id_column, att_first_name_column, att_last_name_column] if col]
                    att_valid_students = att_df[build_student_mask(att_df, att_identifier_columns)]
                    st.info(f"📈 Found **{len(att_valid_students)}** valid students (excluding empty rows)")

                    # Select attendance date columns
                    st.subheader("📅 Attendance Date Columns")

                    # Get available columns (excluding identifier columns)
                    available_columns = [col for col in att_df.columns if col not in att_identifier_columns]

                    # Auto-detect date columns
                    auto_detected_dates = sort_date_columns(detect_date_columns(available_columns))

                    if auto_detected_dates:
                        st.success(f"Auto-detected **{len(auto_detected_dates)}** date columns")
                        default_selection = auto_detected_dates
                    else:
                        st.warning("No date columns auto-detected. Please select manually.")
                        default_selection = []

                    st.caption("Columns with date names are auto-selected. Attendance exports are written in chronological order.")

                    attendance_columns = st.multiselect(
                        "Select date columns",
                        options=available_columns,
                        default=default_selection,
                        key="attendance_columns"
                    )

                    if attendance_columns:
                        st.info(f"**{len(attendance_columns)}** attendance date columns selected")

                        st.divider()

                        # Generate Excel
                        st.subheader("📥 Generate Attendance Excel File")

                        if st.button("🚀 Generate Attendance Excel", type="primary", use_container_width=True, key="generate_attendance"):
                            with st.spinner("🔄 Creating attendance Excel file..."):
                                attendance_output = create_attendance_excel(
                                    att_df,
                                    att_id_column,
                                    att_first_name_column,
                                    att_last_name_column,
                                    attendance_columns
                                )

                                st.success("✅ Attendance Excel file generated successfully!")

                                # Summary
                                st.markdown(f"""
                                <div class="success-box">
                                    <h4>✨ Generation Complete!</h4>
                                    <p>Created <strong>{len(att_valid_students)}</strong> individual student sheets</p>
                                    <p>Each sheet contains:</p>
                                    <ul>
                                        <li>Student ID, First Name, and Last Name</li>
                                        <li>Attendance records by date</li>
                                        <li>Color-coded: 🟢 Green (Present) / 🟠 Orange (Absent)</li>
                                        <li>Attendance summary (days present, total days, rate)</li>
                                    </ul>
                                </div>
                                """, unsafe_allow_html=True)

                                # Download button
                                st.download_button(
                                    label="📥 Download Attendance Excel",
                                    data=attendance_output,
                                    file_name="attendance_report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key="download_attendance"
                                )
                    else:
                        st.warning("Please select at least one attendance date column")

                except ImportError as e:
                    st.error(f"❌ {str(e)}")
                    st.info("💡 Run this command in your terminal: `pip install numbers-parser`")
                except Exception as e:
                    st.error(f"❌ Error parsing file: {str(e)}")
                    st.info("💡 Make sure your Numbers file has a table with headers in the first row")

    # Footer
    st.divider()
    st.markdown(
        "<p style='text-align: center; color: #888;'>GradeBook Transfer | Made for educators</p>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
